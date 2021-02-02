# -*- coding: utf-8 -*-


# TODO: Сортировка листов в excel   !!!!!!!
# TODO: Построение графиков средних при построении профилей, чтобы не удалять по несколько раз выбросы !!!!!!!!

# TODO: Изменить тип графика чтобы были видны точки
# TODO: Изменить названия файлов (добавить широты)
# TODO: Изменить названия диаграммы (добавить широты)

"""
Created on Fri Jan 22 14:31:02 2021


ВРЕМЕННОЙ ХОД ОСРЕДНЕННОГО ПАРАМЕТРА С НОВЫМ УДАЛЕНИЕМ ВЫБРОСОВ (ОБЕ ГРАНИЦЫ ЗАДАЮТСЯ)

1. Фильтрует станции по заданным параметрам.
2. Осредняет значения выбранного параметра за год.
3. Строит график временного хода осредненного параметра.

Опционально:
    - выбор исследуемого района;
    - выбор исследуемого параметра (название колонки в загружаемой таблице 'oxig','sal','temp' и т.д.);
    - удаление выбросов;
    - интерполяция по глубине;
    - запись результатов в xlsx файлы;
    - построение карты расперделения станций;
    - построение графика временного хода осредненного параметра.

"""
import os
import numpy as np
import openpyxl
# =============================================================================
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from scipy import stats

# Путь к папке проекта
path_project = 'D:/УЧЕБА/Самообучение/Програмирование/Python_Projects/Zuenko/Zuenko/py/test/'
# path_project = '/media/lenovo/D/УЧЕБА/Самообучение/Програмирование/Python_Projects/Zuenko/Zuenko/py/test/'
# path_project = 'C:/Users/vladimir.matveev/Desktop/Zuenko/Zuenko/py/test/'
# path_project = 'C:/Users/malyg/Desktop/'

# Загружает файл БД по Охотскому морю
path_orig = f'{path_project}refactoring_base_new.csv'
# path_orig = f'{path_project}tested_2.csv'
df_orig = pd.read_csv(path_orig, sep=',')

# Устанавливает координаты района
# min_lat, max_lat = 51, 59
# # min_lat, max_lat = 51.6, 55
# # min_lat, max_lat = 51, 51.5
# min_long, max_long = 152.9, 157

min_lat, max_lat = 50, 55
# min_lat, max_lat = 55, 58

min_long, max_long = 152, 155

min_lvl, max_lvl = 200, 1100
min_years, max_years = 1980, 2020

# Список условий для фильтрации станций
boundary_area = '(@min_lat      <=  lat     <=  @max_lat) and ' \
                '(@min_long     <=  long    <=  @max_long) and' \
                '(@min_lvl      <=  level   <= @max_lvl) and' \
                '(@min_years    <=  Year    <= @max_years)'

# Выбор исследуемой характеристики/параметра
parameter = 'sal'

# Удаление выбросов, True - удаляет, False - не удаляет (нужное вписать)
outliers_removed = False

# Записывает результаты в excel, True- записывает, False - не записывает (нужное вписать)
to_excel = True

# Делает интерполяцию, True - делает, False - не делает (нужное вписать)
make_interpolation = True

if to_excel:
    if make_interpolation:
        filename_inter = 'means_interpolated'
        filename = f'{path_project}{filename_inter}'
        if not os.path.exists(filename):
            os.mkdir(filename)
    else:
        filename_not_inter = 'means_not_interpolated'
        filename = f'{path_project}{filename_not_inter}'
        if not os.path.exists(filename):
            os.mkdir(filename)

    # name_project_files_1 = '/project_files'
    # name_project_files = 'means_project_files/project_files'
    name_project_files = 'project_files'
    filename_3_1 = f'{filename}/{name_project_files}'
    # filename_3 = f'{path_project}{name_project_files}'


if not to_excel:
    name_project_files_1 = '/project_files'
    name_project_files = '/project_files/means_project_files'
    filename_3_1 = f'{path_project}{name_project_files_1}'
    filename_3 = f'{path_project}{name_project_files}'

if not os.path.exists(filename_3_1):
    os.mkdir(filename_3_1)


# Создает карту распределения станций, True- создает, False - не создает (нужное вписать)
create_map = True

# Создает график вертикальных профилей, True- создает, False - не создает (нужное вписать)
create_graph = True

# Границы уровней
dct_1 = {i: i + 99 for i in range(200, 1000, 100)}
dct_2 = {i: i + 199 for i in range(600, 1000, 200)}
dct_std_lvl = {**dct_1, **dct_2}

# Данные из БД по Охотскому морю по исследуемому району
df_area = df_orig.query(boundary_area)
dff_1 = df_area.copy()
dff_1['level'] = np.round(dff_1['level']).astype(int)
df_area = dff_1.copy()


def create_map_levels(df, min_yrs, max_yrs):
    """
    Строит карту распределения станций по заданным ранее условиям
    """
    dff = df.copy()
    dff = df.query("(@min_yrs   <=   Year   <=  @max_yrs)").copy()

    min_lvl_name = int(dff[['level']].min())
    max_lvl_name = int(dff[['level']].max())

    map_center = go.layout.mapbox.Center(lat=53, lon=149)

    # fig_map = px.scatter_mapbox(dff, lon="long", lat="lat", animation_frame="Year",
    #                             size=parameter,
    #                             hover_name=parameter,
    #                             hover_data={"level": True, "zz": True, 'long': True,
    #                                         "lat": True, parameter: False, 'Year': True},
    #                             size_max=20,
    #                             color=parameter,  # Цветовая кодировка в данном случае по горизонту (0 или 1)
    #                             color_continuous_scale=["yellow", "red"],
    #                             zoom=4,
    #                             center=map_center)
    
    # fig_map.update_layout(
    #     mapbox_style="white-bg",
    #     mapbox_layers=[
    #         {
    #             "below": 'traces',
    #             "sourcetype": "raster",
    #             "sourceattribution": "United States Geological Survey",
    #             "source": [
    #                 "https://basemap.nationalmap.gov/arcgis/rest/services/USGSImageryOnly/MapServer/tile/{z}/{y}/{x}"]
    #         }])
    
    # fig_map.write_html(f'{filename_3_1}/map_stations_area_of_south.html', auto_open=True)

    fig_map_all = px.scatter_mapbox(dff, lon="long", lat="lat",
                                    # size=parameter,
                                    hover_name=parameter,
                                    hover_data={"level": True, "zz": True, 'long': True, 'Month': True,
                                                "lat": True, parameter: False, 'Year': True},
                                    # size_max=15,
                                    color=parameter,  # Цветовая кодировка в данном случае по горизонту (0 или 1)
                                    color_continuous_scale=["yellow", "red"],
                                    zoom=4,
                                    center=map_center)

    fig_map_all.update_layout(
        mapbox_style="white-bg",
        mapbox_layers=[
            {
                "below": 'traces',
                "sourcetype": "raster",
                "sourceattribution": "United States Geological Survey",
                "source": [
                    "https://basemap.nationalmap.gov/arcgis/rest/services/USGSImageryOnly/MapServer/tile/{z}/{y}/{x}"]
            }])

    fig_map_all.write_html(f'{filename_3_1}/{min_lvl_name}_{max_lvl_name}_{min_lat}_{max_lat}.html',
                           auto_open=True)


def name_of_month(month):
    """
    Возвращает название месяца согласно номеру
    """
    dct_month = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
                 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}

    return dct_month[month]


def info_stat(df):
    """
    Получает df
    Выводит на экран размах значений параметра и макс и мин значения

    """
    diff = df[parameter].max() - df[parameter].min()
    diff = np.round(diff, decimals=2)
    min_param = df[parameter].min()
    max_param = df[parameter].max()

    print('========================================================================')
    return f'Размах = {diff},       min_{parameter} = {min_param},      max_{parameter} = {max_param}\n'


def scatter_new(df, lvl):
    """
    Выводит на экран облако точек из полученных данных
    """
    # Добавление в df новых столбцов, 
    # исключительно для косметического эффекта отображения инфо на графике
    dff = df.copy(deep=True)
    dff['x'] = [i for i in range(dff[parameter].count())]
    dff['z_old'] = np.abs(dff['z_score'])
    std = dff[parameter].describe()['std']
    new_z = stats.zscore(dff[parameter])

    dff['std'] = std.round(2)

    dff['new_z'] = new_z.round(2)
    # dff['new_z'] = np.abs( dff['new_z'])

    fig = px.scatter(dff, x='x', y=parameter,
                     hover_name='z_score',
                     color=np.abs(dff['new_z']),
                     size=np.abs(dff['new_z']),

                     hover_data={parameter: True,
                                 'Year': True,
                                 'Month': True,
                                 'Stations': True,
                                 'x': False,
                                 # 'z_score_new':True,
                                 'temp': False,
                                 'sal': False,
                                 'new_z': True,
                                 'z_old': False
                                 },

                     # labels = {'z_score_new':'std'}
                     )

    number_of_month = int(dff['Month'].unique())
    min_year_for_title = min(dff['Year'].unique())
    max_year_for_title = max(dff['Year'].unique())
    title_1 = f'c {min_year_for_title} по {max_year_for_title}'
    title_2 = name_of_month(number_of_month)
    title_3 = lvl
    title_4 = info_stat(dff)
    title = f'{title_1}    {title_2}    {title_3}m    {title_4}'

    fig.update_layout(
        title={
            'text': title,
            'y': 1,
            'x': 0.5,
            'xanchor': 'center',
            'yanchor': 'top'},
        xaxis_title='Порядковый номер',
        yaxis_title="Концентрация растворенного кислорода, мл/л",

        title_font_family='Rockwell ',
        title_font_color="black",
        title_font_size=25)

    fig.write_html(f'{filename_3_1}/{lvl}_m.html', auto_open=True)


def z_score(df, lvl):
    """
    Выдает данные, не превышающие заданый уровень z-стандартизации
    """
    # df_1 = df[[parameter]].copy(deep=True)
    # df_Z = df_1[(np.abs(stats.zscore(df_1)) < lvl).all(axis=1)]
    # ix_keep = df_Z.index
    # df_keep_1 = df.loc[ix_keep]
    # return df_keep_1

    df_1 = df[[parameter]].copy(deep=True)
    # df_Z = df_1[(np.abs(stats.zscore(df_1)) < lvl).all(axis=1)]
    nums = lvl
    df_2 = df_1.copy(deep=True)
    df_2['z_score'] = stats.zscore(df_1)

    if len(nums) == 2:
        for i in nums:
            if i < 0:
                down = i
            elif i > 0:
                up = i

        df_Z = df_2.query('@down <=  z_score <= @up')
        # print(df_Z)

    elif len(nums) == 1:
        for i in nums:
            if i < 0:
                down = i
                df_Z = df_2.query('@down <=  z_score')

            elif i > 0:
                up = i
                df_Z = df_2.query('z_score <= @up')

    ix_keep = df_Z.index

    df_keep_1 = df.loc[ix_keep]

    return df_keep_1


def clean_outliers(df, lvls):
    """
    Ручное удаление выбросов при помощи z-стандартизации.
    \n Если размах в выборке больше заданного:
        \n - показывает карту распеределения станций этой выборки;
        \n - показывает облако точек выборки, также данные о кол-ве стд отклоений;
        \n - предлагает на выбор три варианта: ( " ' " - данный символ не вводить)
            \n 1. Если всё устраивает, то набрать 'ok' и нажать Enter ->
            \n 2. Если не устраивает, то можно набрать границу кол-ва стд откл.:
                 '1.5' - верхняя граница;
                 '-1' - нижняя граница;
                 '1.5 -1' (через пробел) - верхняя и нижняя граница (порядок не важен, важен знак);
                 '2 '(после цифры пробел) - граница с двух сторон == ('2 -2').
            \n 3. 'map' - Если нужна карта станций текущих точек (после удаления выбросов в этой выборке).

   \n Передает дальше общую выборку без выбросов.
    """
    df_old = df.copy()
    dff_concat = pd.DataFrame()
    for month in sorted(df_old.Month.unique()):
        df_old_month = df_old.query('Month == @month')
        # print(df_old_month)

        for k, v in lvls.items():
            min_lvl, max_lvl = k, v
            df_old_1 = df_old_month.query('@min_lvl    <=  level   <=  @max_lvl')

            diff_1 = df_old_1[parameter].max() - df_old_1[parameter].min()
            print(np.round(diff_1, decimals=2))

            # Если размах больше 4, то начинается цикл удаления выбросов
            if diff_1 >= 1:

                df_area_2 = df_old_1[[parameter]].fillna(value=-5.0)
                df_area_1 = df_old_1.copy()
                df_area_1[parameter] = df_area_2

                # Расчет z-значений и добавлений их в df
                df_area_1['z_score'] = stats.zscore(df_area_1[[parameter]]).round(2)

                # Создает крату станций
                create_map_levels(df_area_1, min_years, max_years)

                # Построение облака точек исходных данных
                scatter_new(df_area_1, min_lvl)

                # Показывает размах, макс и мин значения исходных данных
                print(info_stat(df_area_1[[parameter]]))

                # Уровень для z-стандартизации по умолчанию
                dflt_lvl = [100]

                # Новая таблица данных, без изменений, на тот случай, если сразу ok
                df_keep = z_score(df_area_1, dflt_lvl)

                # Цикл удаления выбросов, пока не будет введено оk
                while True:
                    try:
                        mode = input('Введите уровень затем Enter,\nесли устраивает введите ok, затем Enter:\n')
                        if mode == 'ok':
                            dff_concat = pd.concat([dff_concat, df_keep])
                            break
                        elif mode == 'map':
                            create_map_levels(df_keep, min_years, max_years)

                        else:
                            mode_1 = mode.split(' ')
                            level = []
                            for i in mode_1:
                                if i != '':
                                    level.append(float(i))
                                else:
                                    level.append(-level[0])

                            # Расчет данных, не превышающих заданный в ручную уровень z-стандартизации
                            df_keep = z_score(df_area_1, level)

                            # Показывает размах, макс и мин значения расчитанных данных
                            print(info_stat(df_keep))

                            # Построение облака точек изменненых данных
                            # print(df_keep[['level']].describe())
                            # create_map_levels(df_keep, min_years, max_years)
                            scatter_new(df_keep, min_lvl)
                        continue
                    except (IndexError, ValueError):
                        print("\nПопробуйте, ввести уровень еще раз")

            else:
                dff_concat = pd.concat([dff_concat, df_old_1])

    # dff_concat.to_csv(f'{path_project}_dff_concat.csv', index=False)
    return dff_concat


def excel(df, year, name, mode):
    """
    Записывает данные в xlsx
    \n df - данные для записи
    \n year - год, который используется как имя листа
    \n name - имя файла xlsx
    \n mode - режим записи, w-перезапись, a-добавление
    """

    df_to_xlsx = df.copy()
    name_sheet = year
    name_xlsx = name
    mode_xlsx = mode

    # Путь записи файла
    path_xlsx = f'{path_project}{name_xlsx}.xlsx'

    with pd.ExcelWriter(path_xlsx, engine='openpyxl', mode=mode_xlsx) as writer:
        df_to_xlsx.to_excel(writer, index=False, sheet_name=f'{name_sheet}')

    # Удаляет ненужный пустой лист, который был нужен для создания xlsx
    if mode == 'a':
        wb = openpyxl.load_workbook(path_xlsx)
        sheet = wb.sheetnames

        if '1' in sheet:
            pfd = wb['1']
            wb.remove(pfd)
            wb.save(path_xlsx)


def create_empty_xlsx_files():
    """
    Создает пустые файлы xlsx для записи в них результатов
    В каждом файле будет создан лист с именем 1
    """
    df_to_excel = pd.DataFrame()

    lst_file_names = []
    path_directory = filename_inter

    if not make_interpolation:
        path_directory = filename_not_inter

    for i in range(1, 3):

        if i == 1:
            dct_lvl = dct_1
        else:
            dct_lvl = dct_2

        for k, v in dct_lvl.items():
            path_file = f'{path_directory}/{k}_{v + 1}'
            lst_file_names.append(path_file)

    lst_file_names.append(f'{path_directory}/result_{path_directory}')

    for f_name in lst_file_names:
        excel(df_to_excel, 1, f_name, 'w')


def interpolation(df, all_cols):
    """
    Производит линейную интерполяцию \n
    Если all_cols == True, интерполирует все столбцы, иначе все, кроме последнего
    """
    df = df.copy()
    max_lvl = int(df['level'].max())
    df_concat = pd.DataFrame(data={'level': df['level']}, index=[i for i in range(0, max_lvl, 1)])

    if all_cols:
        range_cols = df.columns[1:]
    else:
        range_cols = df.columns[1:-1]

    for nst in range_cols:
        df1 = df[['level', nst]].copy()
        df_inter = pd.DataFrame(data={'level': [i for i in range(0, max_lvl, 1)]},
                                index=[i for i in range(0, max_lvl, 1)])

        df_inter = pd.merge(df_inter, df1[['level', nst]], how='outer', on='level')

        df_inter = df_inter.interpolate(limit_direction='backward', limit_area='inside')
        df_new = pd.merge(df1['level'], df_inter, how='inner', on='level')
        df_concat = pd.concat([df_concat, df_new[nst]], axis=1)
        # df_concat = df_concat.reset_index(drop=True)
        df_concat = df_concat.dropna(how='all')
        df_concat = df_concat.round(2)

    # print('interpolation')    
    # print(df_concat)

    return df_concat


def mean_for_nst_year_lvl(df, min_lvl, max_lvl):
    """
   
    1. Объединяет значения со всех станций за определенный год;\n
    2. Расчитывает среднее значение за год;\n
    3. Объединяет средние за год;\n
    4. Расчитывает среднее среднего;\n
    5. Объединяет средние за год и среднее среднего.\n
    6. Объединяет все средние средних за год.\n
    7. Создает таблицу (Года, Средние значения, с пробелами)
    
    df - данные для обработки\n
    min_lvl - верхняя граница слоя\n
    max_lvl - нижняя граница слоя\n

    """
    min_lvl = min_lvl
    max_lvl = max_lvl

    df = df.copy()

    area_lvl = ('@min_lvl  <=  level  <=  @max_lvl')

    df = df.query(area_lvl)

    df_for_mn_yr_1_lvl = pd.DataFrame()

    path_to_xlsx_all_nst_and_year = f'{filename_inter}/{min_lvl}_{max_lvl + 1}'
    path_to_xlsx_result = f'{filename_inter}/result_{filename_inter}'

    if not make_interpolation:
        path_to_xlsx_all_nst_and_year = f'{filename_not_inter}/{min_lvl}_{max_lvl + 1}'
        path_to_xlsx_result = f'{filename_not_inter}/result_{filename_not_inter}'

    for year in sorted(df['Year'].unique()):

        dff = df.query('Year == @year').copy()

        df_all_nst_for_year = pd.DataFrame(columns=['level'])

        # =============================================================================
        #       Объединяет значения со всех станций за год
        # =============================================================================

        for nst in dff['Stations'].unique():
            new_df = dff.query('Stations == @nst')[['level', parameter]]
            new_df = new_df.rename(columns={parameter: f'nst_{nst}'})

            df_all_nst_for_year = pd.merge(df_all_nst_for_year, new_df, on='level', how="outer")
            df_all_nst_for_year = df_all_nst_for_year.sort_values(by='level')
            df_all_nst_for_year = df_all_nst_for_year.reset_index(drop=True)

        # =============================================================================
        #       Расчет средних значений за год по объединенным значениям со станций
        # =============================================================================

        if make_interpolation:
            df_all_nst_for_year = interpolation(df_all_nst_for_year, True)
            df_all_nst_for_year = df_all_nst_for_year.reset_index(drop=True)

        # Расчитывает среднее значение по всем уровням за год
        mean_for_the_year = df_all_nst_for_year.iloc[:, 1:].agg("mean", axis="columns")
        mean_for_the_year = mean_for_the_year.round(2)

        lst_lvls = df_all_nst_for_year['level'].round()

        # Создает таблицу с уровнями и средним значением за год по этим уровням
        df_mean_for_one_year = pd.DataFrame(data={'level': lst_lvls,
                                                  f'mean_{year}': mean_for_the_year},
                                            index=[i for i in range(len(df_all_nst_for_year['level']))])

        # =============================================================================
        #       Соединяет значения со всех станций со средним за год
        # =============================================================================

        df_all_nst_and_mean_year = pd.merge(df_all_nst_for_year, df_mean_for_one_year, how='inner', on='level')
        print(f'{year}')
        # print(df_all_nst_and_mean_year)
        # =============================================================================
        #        Расчитывает среднее для уровня и соединяет его со значениями станций и среднего за год
        # =============================================================================

        # ЭТО СРЕДНЕЕ_2
        mn_yr_1_lvl = pd.Series(data=df_all_nst_and_mean_year[f'mean_{year}'].mean(), index=[0], name=f'{year}')

        mn_yr_1_lvl = mn_yr_1_lvl.round(2)

        df_all_nst_and_mean_year = pd.concat([df_all_nst_and_mean_year, mn_yr_1_lvl], axis=1)
        df_all_nst_and_mean_year = df_all_nst_and_mean_year.round(2)

        #  Соединеят все средние_2
        df_for_mn_yr_1_lvl = pd.concat([df_for_mn_yr_1_lvl, mn_yr_1_lvl], axis=1)

        if to_excel:
            excel(df_all_nst_and_mean_year, year, path_to_xlsx_all_nst_and_year, 'a')

    #  Переворачивает объедиенные средние_2
    only_mn_yr_1_lvl = df_for_mn_yr_1_lvl.stack()

    only_mn_yr_1_lvl = only_mn_yr_1_lvl.reset_index(drop=True)

    lst_yr = df['Year'].unique()

    # Создает таблицу столбцы года и уровень
    df_only_mn_yr_1_lvl = pd.DataFrame(data={'Year': lst_yr,
                                             f'{min_lvl}_{max_lvl}': only_mn_yr_1_lvl
                                             },
                                       index=[i for i in range(len(lst_yr))])

    lst_yr_nan = [i for i in range(min_years, max_years + 1)]

    df_yr_nan = pd.DataFrame(data={'Year': lst_yr_nan},
                             index=[i for i in range(len(lst_yr_nan))])

    df_yr_nan = pd.merge(df_yr_nan, df_only_mn_yr_1_lvl, on='Year', how='outer')

    if to_excel:
        excel(df_yr_nan, 'all_years', path_to_xlsx_all_nst_and_year, 'a')

    return df_yr_nan


def graph_excel(lst_year, title_excel, yaxis_title_excel):
    """
    В итоговом файле xlsx, создает диаграмму\n
    lst_year - для количества столбцов, по которым строится график\n
    title_excel - название диаграммы\n
    yaxis_title_excel - название оси Y\n
    """
    from openpyxl import Workbook
    from openpyxl.chart import (
        LineChart,
        ScatterChart,
        Reference,
        Series,
    )

    wb = openpyxl.load_workbook(f'{path_project}{filename_inter}/result_{filename_inter}.xlsx')

    ws = wb['all']

    chart = ScatterChart()
    chart.title = title_excel
    # chart.style = 2
    chart.x_axis.title = 'Года'
    chart.y_axis.title = yaxis_title_excel

    # chart.x_axis.scaling.min = 0
    # chart.y_axis.scaling.min = 1
    # # chart.x_axis.scaling.max = 11
    # chart.y_axis.scaling.max = 2.7

    num = 2
    max_rows = len(lst_year)
    for i in range(1, 3):

        if i == 1:
            dct_lvl = dct_1
        else:
            dct_lvl = dct_2

        for k, v in dct_lvl.items():
            xvalues = Reference(ws, min_col=1, min_row=2, max_row=1 + max_rows)
            values = Reference(ws, min_col=num, min_row=1, max_row=1 + max_rows)
            # print('values')
            # print(values)
            num += 1

            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)

    ws.add_chart(chart, "K02")

    wb.save(f'{path_project}{filename_inter}/result_{filename_inter}.xlsx')


def graph_profile_of_means():
    """
    Создает график средних годовых значений выбранного параметра
    """

    df = df_area.copy()
    fig_graph = go.Figure()
    # print(df)
    # Карта распределения станций
    if create_map:
        create_map_levels(df, min_years, max_years)

    # if outliers_removed:
    #     df = clean_outliers(df)

    if to_excel:
        # Создает пустые файлы xlsx
        create_empty_xlsx_files()

    lst_years = [i for i in range(min_years, max_years + 1)]

    df_result = pd.DataFrame(data={'Year': lst_years},
                             index=[i for i in range(len(lst_years))])

    path_to_xlsx_result = f'{filename_inter}/result_{filename_inter}'

    if not make_interpolation:
        path_to_xlsx_result = f'{filename_not_inter}/result_{filename_not_inter}'

    for i in range(1, 3):

        if i == 1:
            dct_lvl = dct_1
        else:
            dct_lvl = dct_2

        if outliers_removed:
            df = clean_outliers(df, dct_lvl)

        for k, v in dct_lvl.items():
            result = mean_for_nst_year_lvl(df, k, v)
            # graph_excel(result,k,v)
            df_result = pd.merge(df_result, result, on='Year', how='outer')

            x = result['Year']
            y = result[f'{k}_{v}']

            fig_graph.add_trace(go.Scatter(x=x, y=y, name=f'{k}-{v}'))

    if to_excel:
        excel(df_result, 'all', path_to_xlsx_result, 'a')

    if create_graph:
        # Подписи к графику
        min_lvl = min([i for i in dct_1.keys()])
        max_lvl = max([i for i in dct_1.values()])

        # Подписи к графику
        if parameter == 'oxig':
            title = 'Средние показатели растворенного кислорода'
            y_axis_title = 'Концентрация растворенного кислорода, мл/л'

        elif parameter == 'sal':
            title = 'Средние показатели солености'
            y_axis_title = 'Соленость, е.п.с.'

        elif parameter == 'temp':
            title = 'Средние показатели температуры'
            y_axis_title = 'Температура, С'

        fig_graph.update_layout(
            title={
                'text': f"{title} на глубинах {min_lvl} - {max_lvl + 1} м ",
                'y': 1,
                'x': 0.5,
                'xanchor': 'center',
                'yanchor': 'top'},

            xaxis_title="Год",
            yaxis_title=f'{y_axis_title}',
            title_font_family="Arial, bold",
            title_font_color="black",
            title_font_size=30)

        fig_graph.write_html(f'{filename_3_1}/graph_mean.html', auto_open=True)

        # =============================================================================
        #       Создает график в Excel
        # =============================================================================
        if to_excel:
            graph_excel(lst_years, title, y_axis_title)


if __name__ == "__main__":
    graph_profile_of_means()
