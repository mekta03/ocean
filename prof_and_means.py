# -*- coding: utf-8 -*-
"""
Created on Fri Jan 22 13:40:32 2021

ВЕРТИКАЛЬНЫЕ ПРОФИЛИ РАСПРЕДЕЛЕНИЯ ВЫБРАННОГО ПАРАМЕТРА

1. Фильтрует станции по заданным параметрам.
2. Осредняет и приводит к стандартным горизонтам выбранный параметр.
3. Строит вертикальные профили выбранного параметра.

Опционально:
    - выбор исследуемого района;
    - выбор исследуемого параметра (название колонки в загружаемой таблице 'oxig','sal','temp' и т.д.);
    - удаление выбросов;
    - интерполяция по глубине;
    - запись результатов в xlsx файлы;
    - строит график в excel;
    - построение карты расперделения станций;
    - построение графика вертикальных профилей.

"""
# =============================================================================
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from scipy import stats
import os

# Путь к папке проекта
# path_project = 'C:/Users/vladimir.matveev/Desktop/Zuenko/Zuenko/py/test/'
path_project = 'D:/УЧЕБА/Самообучение/Програмирование/Python_Projects/Zuenko/Zuenko/py/test/'
# path_project = '/media/lenovo/D/УЧЕБА/Самообучение/Програмирование/Python_Projects/Zuenko/Zuenko/py/test/'


# Загружает файл БД по Охотскому морю
path_orig = f'{path_project}refactoring_base_new.csv'

# Загружает файл БД по Охотскому морю
df_orig = pd.read_csv(path_orig, sep=',')

# Устанавливает координаты района
# min_lat, max_lat = 55, 58.5
# min_lat, max_lat = 51.6, 55
# min_lat, max_lat = 51, 51.5
# min_long, max_long = 152.9, 157

# Новые границы
# min_lat, max_lat = 50, 55
min_lat, max_lat = 55, 58
# min_lat, max_lat = 50, 58
min_long, max_long = 152, 155

min_zz = 300
min_years, max_years = 1980, 2020

# Список условий для фильтрации станций
boundary_area = '(@min_lat      <=  lat     <=  @max_lat) and ' \
                '(@min_long     <=  long    <=  @max_long) and' \
                '(@min_zz       <=  zz )'

# Выбор исследуемой характеристики/параметра
parameter = 'sal'

# Удаление выбросов, True - удаляет, False - не удаляет (нужное вписать)
outliers_removed = False

# Записывает результаты в excel, True- записывает, False - не записывает (нужное вписать)
to_excel = True

# Делает интерполяцию, True - делает, False - не делает (нужное вписать)
make_interpolation = True

if to_excel:
    if make_interpolation:
        filename_prof_inter = f'profile_interpolated_{min_lat}-{max_lat}'
        filename_prof = f'{path_project}{filename_prof_inter}'
        if not os.path.exists(filename_prof):
            os.mkdir(filename_prof)
    else:
        filename_prof_not_inter = f'profile_not_interpolated_{min_lat}-{max_lat}'
        filename_prof = f'{path_project}{filename_prof_not_inter}'
        if not os.path.exists(filename_prof):
            os.mkdir(filename_prof)

    # name_project_prof_files = 'project_files/profile_project_files'
    name_project_prof_files = 'project_files'
    filename_prof_1 = f'{filename_prof}/{name_project_prof_files}'
    # filename_3 = f'{path_project}{name_project_prof_files}'

if not to_excel:
    name_project_prof_files = 'project_files/profile_project_files'
    name_project_prof_files_1 = 'project_files'
    filename_prof_1 = f'{path_project}{name_project_prof_files_1}'
    filename_prof_1_1 = f'{path_project}{name_project_prof_files}'

if not os.path.exists(filename_prof_1):
    os.mkdir(filename_prof_1)
    # os.mkdir(filename_3)

# Имена файлов xlsx
all_dec_inter = f'all_dec_inter_{min_lat}-{max_lat}'
all_dec_not_inter = f'all_dec_not_inter_{min_lat}-{max_lat}'
rslt_std_inter = f'rslt_std_all_dec_inter_{min_lat}-{max_lat}'
rslt_std_not_inter = f'rslt_std_all_dec_not_inter_{min_lat}-{max_lat}'

# Создает карту распределения станций, True- создает, False - не создает (нужное вписать)
create_map = True

# Создает график вертикальных профилей, True- создает, False - не создает (нужное вписать)
create_graph = True

# Границы уровней
dct_1 = {i: i + 9 for i in range(0, 31, 10)}
dct_2 = {i: i + 19 for i in range(30, 31)}
dct_3 = {i: i + 49 for i in range(50, 251, 50)}
dct_4 = {i: i + 99 for i in range(300, 1001, 100)}
dct_std_lvl = {**dct_1, **dct_2, **dct_3, **dct_4}

# Данные из БД по Охотскому морю по исследуемому району
df_area = df_orig.query(boundary_area)
dff_1 = df_area.copy()
dff_1['level'] = np.round(dff_1['level']).astype(int)
df_area = dff_1.copy()


def create_map_levels(df, min_yrs, max_yrs):
    """
    Строит карту распределения станций по заданным ранее условиям
    """
    dff = df.copy()
    dff = df.query("(@min_yrs   <=   Year   <=  @max_yrs)").copy()

    min_lvl_name = int(dff[['level']].min())
    max_lvl_name = int(dff[['level']].max())

    map_center = go.layout.mapbox.Center(lat=53, lon=149)
    #
    # fig_map = px.scatter_mapbox(dff, lon="long", lat="lat", animation_frame="Year",
    #                             size=parameter,
    #                             hover_name=parameter,
    #                             hover_data={"level": True, "zz": True, 'long': True, 'Stations':True,
    #                                         "lat": True, parameter: False, 'Year': False},
    #                             size_max=20,
    #                             color=parameter,  # Цветовая кодировка в данном случае по горизонту (0 или 1)
    #                             color_continuous_scale=["yellow", "red"],
    #                             zoom=4,
    #                             center=map_center)
    #
    # fig_map.update_layout(
    #     mapbox_style="white-bg",
    #     mapbox_layers=[
    #         {
    #             "below": 'traces',
    #             "sourcetype": "raster",
    #             "sourceattribution": "United States Geological Survey",
    #             "source": [
    #
    #                 "https://basemap.nationalmap.gov/arcgis/rest/services/USGSImageryOnly/MapServer/tile/{z}/{y}/{x}"]
    #         }])
    #
    # fig_map.write_html(f'{path_project}{name_project_prof_files}/map_stations_area_of_south.html', auto_open=True)
    #

    fig_map_all = px.scatter_mapbox(dff, lon="long", lat="lat",
                                    # size=parameter,
                                    hover_name=parameter,
                                    hover_data={"level": True, "zz": True, 'long': True, "lat": True, 'Stations': True,
                                                parameter: False, 'Year': True},
                                    size_max=20,
                                    color=parameter,  # Цветовая кодировка в данном случае по горизонту (0 или 1)
                                    color_continuous_scale=["yellow", "red"],
                                    zoom=4,
                                    center=map_center)

    fig_map_all.update_layout(
        mapbox_style="white-bg",
        mapbox_layers=[
            {
                "below": 'traces',
                "sourcetype": "raster",
                "sourceattribution": "United States Geological Survey",
                "source": [
                    "https://basemap.nationalmap.gov/arcgis/rest/services/USGSImageryOnly/MapServer/tile/{z}/{y}/{x}"]
            }])

    fig_map_all.write_html(
        f'{filename_prof}/{name_project_prof_files}/{min_lvl_name}_{max_lvl_name}m__coord_{min_lat}_{max_lat}.html',
        auto_open=True)


def name_of_month(month):
    """
    Возвращает название месяца согласно номеру
    """
    dct_month = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
                 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}

    return dct_month[month]


def info_stat(df):
    """
    Получает df
    Выводит на экран размах значений параметра и макс и мин значения

    """
    diff = df[parameter].max() - df[parameter].min()
    diff = np.round(diff, decimals=2)
    min_param = df[parameter].min()
    max_param = df[parameter].max()

    print('==========================================================')
    return f'| Размах = {diff} | min_{parameter} = {min_param} | max_{parameter} = {max_param}\n'


def scatter_new(df, lvl):
    """
    Выводит на экран облако точек из полученных данных
    """
    # Добавление в df новых столбцов, 
    # исключительно для косметического эффекта отображения инфо на графике
    dff = df.copy(deep=True)
    dff['x'] = [i for i in range(dff[parameter].count())]
    dff['z_old'] = np.abs(dff['z_score'])
    std = dff[parameter].describe()['std']
    new_z = stats.zscore(dff[parameter])

    dff['std'] = std.round(2)

    dff['new_z'] = new_z.round(2)
    # dff['new_z'] = np.abs( dff['new_z'])

    fig = px.scatter(dff, x='x', y=parameter,
                     hover_name='z_score',
                     color=np.abs(dff['new_z']),
                     size=np.abs(dff['new_z']),
                     color_continuous_scale=["yellow", "red"],
                     hover_data={parameter: True,
                                 'Year': True,
                                 'Month': True,
                                 'Stations': True,
                                 'x': False,
                                 # 'z_score_new':True,
                                 'temp': False,
                                 'sal': False,
                                 'new_z': True,
                                 'z_old': False
                                 },

                     # labels = {'z_score_new':'std'}
                     )

    # print('dff')
    # print(dff['Year'])
    number_of_month = int(dff['Month'].unique())
    min_year_for_title = min(dff['Year'].unique())
    max_year_for_title = max(dff['Year'].unique())
    title_1 = f'c {min_year_for_title} по {max_year_for_title}'
    title_2 = name_of_month(number_of_month)
    title_3 = lvl
    title_4 = info_stat(dff)

    title = f'{title_1}    {title_2}    {title_3}m    {title_4}'
    fig.update_layout(
        title={
            'text': title,
            'y': 1,
            'x': 0.5,
            'xanchor': 'center',
            'yanchor': 'top'},
        xaxis_title='Порядковый номер',
        yaxis_title="Концентрация растворенного кислорода, мл/л",

        title_font_family='Rockwell ',
        title_font_color="black",
        title_font_size=25)

    fig.write_html(f'{filename_prof}/{name_project_prof_files}/{lvl}_m.html', auto_open=True)


def z_score(df, lvl):
    """
    Выдает данные, не превышающие заданый уровень z-стандартизации
    """

    df_1 = df[[parameter]].copy(deep=True)
    nums = lvl
    df_2 = df_1.copy(deep=True)
    df_2['z_score'] = stats.zscore(df_1)

    if len(nums) == 2:
        for i in nums:
            if i < 0:
                down = i
            elif i > 0:
                up = i

        df_z = df_2.query('@down <=  z_score <= @up')
        # print(df_z)

    elif len(nums) == 1:
        for i in nums:
            if i < 0:
                down = i
                df_z = df_2.query('@down <=  z_score')

            elif i > 0:
                up = i
                df_z = df_2.query('z_score <= @up')

    ix_keep = df_z.index

    df_keep_1 = df.loc[ix_keep]

    return df_keep_1


def clean_outliers(df, min_yrs, max_yrs):
    """
    Ручное удаление выбросов при помощи z-стандартизации.
    \n Если размах в выборке больше заданного:
        \n - показывает карту распеределения станций этой выборки;
        \n - показывает облако точек выборки, также данные о кол-ве стд отклоений;
        \n - предлагает на выбор три варианта: ( " ' " - данный символ не вводить)
            \n 1. Если всё устраивает, то набрать 'ok' и нажать Enter ->
            \n 2. Если не устраивает, то можно набрать границу кол-ва стд откл.:
                 '1.5' - верхняя граница;
                 '-1' - нижняя граница;
                 '1.5 -1' (через пробел) - верхняя и нижняя граница (порядок не важен, важен знак);
                 '2 '(после цифры пробел) - граница с двух сторон == ('2 -2').
            \n 3. 'map' - Если нужна карта станций текущих точек (после удаления выбросов в этой выборке).

   \n Передает дальше общую выборку без выбросов.
    """
    df_old = df.copy()
    dff_concat = pd.DataFrame()
    for month in sorted(df_old.Month.unique()):
        df_old_month = df_old.query('Month == @month')
        # print(df_old_month)
        for k, v in dct_std_lvl.items():
            min_lvl, max_lvl = k, v
            df_old_1 = df_old_month.query('@min_lvl    <=  level   <=  @max_lvl')

            diff_1 = df_old_1[parameter].max() - df_old_1[parameter].min()
            print(np.round(diff_1, decimals=2))
            if diff_1 >= 2:

                df_area_2 = df_old_1[[parameter]].fillna(value=-5.0)
                df_area_1 = df_old_1.copy()
                df_area_1[parameter] = df_area_2

                # Расчет z-значений и добавлений их в df
                df_area_1['z_score'] = stats.zscore(df_area_1[[parameter]]).round(2)

                # Создает крату станций
                create_map_levels(df_area_1, min_yrs, max_yrs)

                # Построение облака точек исходных данных
                scatter_new(df_area_1, min_lvl)

                # Показывает размах, макс и мин значения исходных данных
                print(info_stat(df_area_1[[parameter]]))

                # Уровень для z-стандартизации по умолчанию
                dflt_lvl = [100]

                # Новая таблица данных, без изменений, на тот случай, если сразу ok
                df_keep = z_score(df_area_1, dflt_lvl)

                while True:
                    try:
                        mode = input('Введите уровень затем Enter,\n если устраивает введите ok, затем Enter:\n')
                        if mode == 'ok':
                            dff_concat = pd.concat([dff_concat, df_keep])
                            break
                        elif mode == 'map':
                            create_map_levels(df_keep, min_years, max_years)
                        else:
                            mode_1 = mode.split(' ')
                            level = []
                            for i in mode_1:
                                if i != '':
                                    level.append(float(i))
                                else:
                                    level.append(-level[0])

                            # Расчет данных, не превышающих заданный в ручную уровень z-стандартизации
                            df_keep = z_score(df_area_1, level)

                            # Показывает размах, макс и мин значения расчитанных данных
                            print(info_stat(df_keep))

                            # Построение облака точек изменненых данных
                            # create_map_levels(df_keep, min_yrs, max_yrs)
                            scatter_new(df_keep, min_lvl)
                        continue
                    except (IndexError, ValueError):
                        print("\nПопробуйте, ввести уровень еще раз")

            else:
                dff_concat = pd.concat([dff_concat, df_old_1])
    # dff_concat.to_csv(f'{path_project}concat.csv', index=False)
    return dff_concat


def excel(df, year, name, mode):
    """
    Записывает данные в xlsx
    \n df - данные для записи
    \n year - год, который используется как имя листа
    \n name - имя файла xlsx
    \n mode - режим записи, w-перезапись, a-добавление
    """

    df_to_xlsx = df.copy()
    name_sheet = year
    name_xlsx = name
    mode_xlsx = mode

    # Путь записи файла
    path_xlsx = f'{path_project}{name_xlsx}.xlsx'

    with pd.ExcelWriter(path_xlsx, engine='openpyxl', mode=mode_xlsx) as writer:
        df_to_xlsx.to_excel(writer, index=False, sheet_name=f'{name_sheet}')

    # Удаляет ненужный пустой лист, который был нужен для создания xlsx
    if mode == 'a':
        wb = openpyxl.load_workbook(path_xlsx)
        sheet = wb.sheetnames

        if '1' in sheet:
            pfd = wb['1']
            wb.remove(pfd)
            wb.save(path_xlsx)


def create_empty_xlsx_files(dct_years):
    """
    Создает пустые файлы xlsx для записи в них результатов
    В каждом файле будет создан лист с именем 1
    """
    df_to_excel = pd.DataFrame()

    if make_interpolation:
        inter_all_dec = f'{filename_prof_inter}/{all_dec_inter}'
        # std_lvl_decade = f'{filename_prof_inter}/std_lvl_all_decade_interpolated'
        std_lvl_decade = f'{filename_prof_inter}/{rslt_std_inter}'
        names_file = [inter_all_dec, std_lvl_decade]

    else:
        not_inter_all_dec = f'{filename_prof_not_inter}/{all_dec_not_inter}'
        std_lvl_decade = f'{filename_prof_not_inter}/{rslt_std_not_inter}'
        names_file = [not_inter_all_dec, std_lvl_decade]

    for name1 in names_file:
        excel(df_to_excel, 1, name1, 'w')

    for k, v in dct_years.items():

        if make_interpolation:
            inter = f'{filename_prof_inter}/{filename_prof_inter}_{k}'
            names_file_1 = inter
        else:
            not_inter = f'{filename_prof_not_inter}/{filename_prof_not_inter}_{k}'
            names_file_1 = not_inter

        for name1 in [names_file_1]:
            excel(df_to_excel, 1, name1, 'w')


def interpolation(df, all_cols):
    """
    Производит линейную интерполяцию \n
    Если all_cols == True, интерполирует все столбцы, иначе все, кроме последнего
    """
    df = df.copy()
    max_lvl = int(df['level'].max())
    df_concat = pd.DataFrame(data={'level': df['level']}, index=[i for i in range(0, max_lvl, 1)])

    if all_cols:
        range_cols = df.columns[1:]
    else:
        range_cols = df.columns[1:-1]

    for nst in range_cols:
        df1 = df[['level', nst]].copy()
        df_inter = pd.DataFrame(data={'level': [i for i in range(0, max_lvl, 1)]},
                                index=[i for i in range(0, max_lvl, 1)])

        df_inter = pd.merge(df_inter, df1[['level', nst]], how='outer', on='level')

        df_inter = df_inter.interpolate(limit_direction='backward', limit_area='inside')
        df_new = pd.merge(df1['level'], df_inter, how='inner', on='level')
        df_concat = pd.concat([df_concat, df_new[nst]], axis=1)
        # df_concat = df_concat.reset_index(drop=True)
        df_concat = df_concat.dropna(how='all')
        df_concat = df_concat.round(2)

    # print('interpolation')    
    # print(df_concat)

    return df_concat


def mean_for_nst_year_decade(df, name_of_decade):
    """
   
    1. Объединяет значения со всех станций за определенный год;\n
    2. Расчитывает среднее значение за год;\n
    3. Объединяет средние за год;\n
    4. Расчитывает среднее за декаду;\n
    5. Объединяет средние за год и среднее за декаду.\n
    
    df - данные для обработки\n
    name_of_decade - какое дестилетие (например 1980-е)\n
    """

    df = df.copy()
    df_mean_for_all_years = pd.DataFrame(columns=["level"])

    decade = name_of_decade

    for year in sorted(df['Year'].unique()):

        dff = df.query('Year == @year').copy()

        df_all_nst_for_year = pd.DataFrame(columns=['level'])

        # =============================================================================
        #       Объединяет значения со всех станций за год
        # =============================================================================
        print(f'{year}')
        for nst in dff['Stations'].unique():
            new_df = dff.query('Stations == @nst')[['level', parameter]]
            new_df = new_df.rename(columns={parameter: f'nst_{nst}'})

            df_all_nst_for_year = pd.merge(df_all_nst_for_year, new_df, on='level', how="outer")
            df_all_nst_for_year = df_all_nst_for_year.sort_values(by='level')
            df_all_nst_for_year = df_all_nst_for_year.reset_index(drop=True)
        
            # print(df_all_nst_for_year)
        # =============================================================================
        #       Расчет средних значений за год по объединенным значениям со станций
        # =============================================================================

        if not make_interpolation:
            path_to_xlsx_all_nst_and_year = f'{filename_prof_not_inter}/{filename_prof_not_inter}_{decade}'
            path_to_xlsx_all_decade = f'{filename_prof_not_inter}/{all_dec_not_inter}'
        else:
            path_to_xlsx_all_nst_and_year = f'{filename_prof_inter}/{filename_prof_inter}_{decade}'
            path_to_xlsx_all_decade = f'{filename_prof_inter}/{all_dec_inter}'
            df_all_nst_for_year = interpolation(df_all_nst_for_year, True)
            df_all_nst_for_year = df_all_nst_for_year.reset_index(drop=True)

        # Расчитывает среднее значение по всем уровням за год
        mean_for_the_year = df_all_nst_for_year.iloc[:, 1:].agg("mean", axis="columns")
        mean_for_the_year = mean_for_the_year.round(2)

        lst_lvls = df_all_nst_for_year['level'].round()

        # Создает таблицу с уровнями и средним значением за год по этим уровням
        df_mean_for_one_year = pd.DataFrame(data={'level': lst_lvls,
                                                  f'mean_{year}': mean_for_the_year},
                                            index=[i for i in range(len(df_all_nst_for_year['level']))])

        # =============================================================================
        #       Соединяет значения со всех станций со средним за год
        # =============================================================================

        df_nst_year = pd.merge(df_all_nst_for_year, df_mean_for_one_year, how='inner', on='level')

        if to_excel:
            excel(df_nst_year, year, path_to_xlsx_all_nst_and_year, 'a')

        # =============================================================================
        #       Соединяет средние значения за каждый год         
        # =============================================================================

        df_mean_for_all_years = pd.merge(df_mean_for_all_years, df_mean_for_one_year,
                                         how='outer', on='level')

        df_mean_for_all_years = df_mean_for_all_years.sort_values(by='level')
        df_mean_for_all_years = df_mean_for_all_years.reset_index(drop=True)

    if make_interpolation:
        df_mean_for_all_years = interpolation(df_mean_for_all_years, True)

    df_mean_for_all_years = df_mean_for_all_years.round(2)

    # =============================================================================
    #   Расчитывает среднее значение за десятилетие
    # =============================================================================
    mean_for_dec = df_mean_for_all_years.iloc[:, 1:].agg("mean", axis="columns")

    # =============================================================================
    #   Добавляет в таблицу со средними значениями за года среднее за десятилетие
    # =============================================================================
    df_mean_for_all_years[f'{parameter}_{decade}s'] = mean_for_dec.round(2)

    if to_excel:
        # Записывает значения за декаду в файл с отдельной декадой
        excel(df_mean_for_all_years, f'{decade}s', path_to_xlsx_all_nst_and_year, 'a')

        # Записывает значения за декаду в файл только с декадами
        excel(df_mean_for_all_years, f'{decade}s', path_to_xlsx_all_decade, 'a')

    return df_mean_for_all_years


def mean_year_decade_to_std_lvl(start_dec, end_dec):
    """
    1. Приводит рассчитанные ср. знач. за года и декады к стандартным горизонтам
    
    [0,10,20,30,50,100...250,300,400...1000]
    
    start_dec - год начала дестилетия
    end_dec - год конца дестилетия
    """
    min_year, max_year = start_dec, end_dec
    df = df_area.query('(@min_year <= Year <= @max_year)').copy()

    # Вариант удаления выбросов
    if outliers_removed:
        df_new = clean_outliers(df, min_year, max_year)
    else:
        df_new = df.copy()
    # Расчитывает ср.знач. за года и десятилетия
    df_means_year_dec = mean_for_nst_year_decade(df_new, min_year)

    df_means_std_level = pd.DataFrame()

    for col_name in df_means_year_dec.columns[1:]:
        lst_lvl_start = []
        lst_lvl_end = []
        lst_means = []

        dff = df_means_year_dec[['level', col_name]]

        for k, v in dct_std_lvl.items():
            min_lvl, max_lvl = k, v

            df_means_year_dec_for_lvl = dff.query('@min_lvl    <=  level   <=  @max_lvl')

            lst_lvl_start.append(min_lvl)
            lst_lvl_end.append(max_lvl)

            mean_of_parameter = df_means_year_dec_for_lvl[col_name].mean()

            lst_means.append(mean_of_parameter)

        # Создает таблицу со ср.знач. привиденными к стд.горизонтам
        df_mean_std = pd.DataFrame(data={col_name: lst_means},
                                   index=[i for i in range(len(lst_lvl_start))])

        df_mean_std = df_mean_std.round(2)

        df_means_std_level['level'] = lst_lvl_start

        # Если в колонке находтся значения декады, то она записывает в отдельную таблицу
        if col_name != f'{parameter}_{min_year}s':
            df_means_std_level = pd.concat([df_means_std_level, df_mean_std[col_name]], axis=1)

        else:
            df_means_std_level = pd.concat([df_means_std_level, df_mean_std[col_name]], axis=1)
            df_mean_decade_std_lvl = pd.DataFrame(data={'level': lst_lvl_start,
                                                        'level_2': lst_lvl_end,
                                                        f'{parameter}_{min_year}s': lst_means
                                                        },
                                                  index=[i for i in range(len(lst_lvl_start))])

            df_mean_decade_std_lvl = df_mean_decade_std_lvl.round(2)

    if to_excel:
        if make_interpolation:
            excel(df_means_std_level, min_year, f'{filename_prof_inter}/{rslt_std_inter}', 'a')
        else:
            excel(df_means_std_level, min_year, f'{filename_prof_not_inter}/{rslt_std_not_inter}',
                  'a')

    return df_mean_decade_std_lvl


def graph_excel(dct_years, title_excel, yaxis_title_excel):
    """
    В итоговом файле xlsx, создает диаграмму\n
    dct_years - для количества столбцов, по которым строится график\n
    title_excel - название диаграммы\n
    yaxis_title_excel - название оси Y\n
    """
    # from openpyxl.styles import Font, Fill
    from openpyxl import Workbook
    from openpyxl.chart import (
        ScatterChart,
        Reference,
        Series)

    num = 2

    if make_interpolation:
        wb = openpyxl.load_workbook(f'{path_project}{filename_prof_inter}/{rslt_std_inter}.xlsx')
    else:
        wb = openpyxl.load_workbook(
            f'{path_project}{filename_prof_not_inter}/{rslt_std_not_inter}.xlsx')

    ws = wb['all_decades']
    # ws.font = Font(size=25)

    chart = ScatterChart()
    chart.title = title_excel
    chart.style = 2
    chart.x_axis.title = yaxis_title_excel
    chart.y_axis.title = 'Глубина, м'

    # chart.x_axis.scaling.min = 0
    # chart.y_axis.scaling.min = 1
    # chart.x_axis.scaling.max = 11
    # chart.y_axis.scaling.max = 2.7

    #========================================================
    # Расчет максимальных и минимальных значений для осей
    df_excel = pd.DataFrame(ws.values)

    num_of_cols = df_excel.iloc[:,1:].shape[1]

    lst_max = []
    lst_min = []

    for i in range(1, 1+num_of_cols):
        max_of_series = df_excel.iloc[1:,i].max()
        min_of_series = df_excel.iloc[1:,i].min()
        lst_max.append(max_of_series)
        lst_min.append(min_of_series)

    max_x = np.round(max(lst_max)+1)
    min_x = np.round(min(lst_min)-1)
    max_y = 1000
    min_y = 0
    #=============================================================

    max_rows = len(dct_std_lvl.keys())

    for k, v in dct_years.items():
        values = Reference(ws, min_col=1, min_row=2, max_row=1 + max_rows)
        xvalues = Reference(ws, min_col=num, min_row=2, max_row=1 + max_rows)
        # print('values')
        # print(values)
        num += 1

        series = Series(values, xvalues, title_from_data=False)

        chart.series.append(series)
        chart.y_axis.scaling.orientation = "maxMin"
        
        # Установление макс и мин значений осей
        chart.y_axis.scaling.max = max_y
        chart.y_axis.scaling.min = min_y
        chart.x_axis.scaling.max = max_x
        chart.x_axis.scaling.min = min_x

    ws.add_chart(chart, "K02")

    if make_interpolation:
        wb.save(f'{path_project}{filename_prof_inter}/{rslt_std_inter}.xlsx')
    else:
        wb.save(f'{path_project}{filename_prof_not_inter}/{rslt_std_not_inter}.xlsx')


def graph_profile_of_means():
    """
    Создает вертикальные профили распределения выбранного параметра по дестилетиям
    """
    # Карта распределения станций
    if create_map:
        create_map_levels(df_area, min_years, max_years)

    fig_graph = go.Figure()

    # Словарь десятилетий (начало дестилетия: конец дестилетия)
    dct_years = {i: i + (9 if i < 2010 else 10) for i in range(min_years, max_years, 10)}

    if to_excel:
        # Создает пустые файлы xlsx
        create_empty_xlsx_files(dct_years)

    df_std_lvl_mean_all_decades = pd.DataFrame()

    for k, v in dct_years.items():
        min_year, max_year = k, v

        df = mean_year_decade_to_std_lvl(min_year, max_year)

        print(f'{min_year}')

        x = df[f'{parameter}_{min_year}s'].round(2)

        y = df['level']

        df_std_lvl_mean_all_decades['level'] = y
        df_std_lvl_mean_all_decades = pd.concat([df_std_lvl_mean_all_decades, x], axis=1)

        name_of_decade = f'{min_year}`s'
        fig_graph.add_trace(go.Scatter(x=x, y=y, name=name_of_decade, line_shape='spline'))
        fig_graph.update_traces(hovertemplate="Кислород: %{x}<br>Глубина: %{y}+")

    if to_excel:
        if make_interpolation:
            file_name_3 = f'{filename_prof_inter}/{rslt_std_inter}'
        else:
            file_name_3 = f'{filename_prof_not_inter}/{rslt_std_not_inter}'
        excel(df_std_lvl_mean_all_decades, 'all_decades', file_name_3, 'a')

    if create_graph:
        fig_graph.update_layout(margin=dict(l=500, r=500, t=50, b=20))
        fig_graph.update_yaxes(autorange="reversed")

        if parameter == 'oxig':
            title = f"Вертикальные профили кислорода с {min_years} по {max_years} гг. ({min_lat}-{max_lat})"
            axis_title = "Концентрация растворенного кислорода, мл/л"
        elif parameter == 'sal':
            title = f"Вертикальные профили солености с {min_years} по {max_years} гг. ({min_lat}-{max_lat})"
            axis_title = "Соленость, е.п.с."
        elif parameter == 'temp':
            title = f"Вертикальные профили температуры с {min_years} по {max_years} гг. ({min_lat}-{max_lat})"
            axis_title = "Температура, С"

        fig_graph.update_layout(
            title={
                'text': title,
                'y': 1,
                'x': 0.5,
                'xanchor': 'center',
                'yanchor': 'top'},
            xaxis_title=axis_title,
            yaxis_title="Глубина, м",

            title_font_family="Arial, bold",
            title_font_color="black",
            title_font_size=35)

        fig_graph.write_html(f'{filename_prof}/{name_project_prof_files}/{min_lat}_{parameter}_profiles.html', auto_open=True)

        # =============================================================================
        #       Создает график в Excel
        # =============================================================================
        if to_excel:
            graph_excel(dct_years, title, axis_title)

    print('\nThe calculation is successfully completed!')


if __name__ == "__main__":
    graph_profile_of_means()
